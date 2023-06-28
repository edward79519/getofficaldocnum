import datetime

from docnum.models import Contract
from django.utils import timezone
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from django.db.models import Value, CharField
from django.db.models.functions import Concat
from pathlib import Path
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from officaldoc.settings import HOST_IP, HOST_PORT


# today localtime
today = timezone.localtime(timezone.now()).date()

# report temp dir
TMP_DIR = Path('static', 'tmp')
if not TMP_DIR.exists():
    Path.mkdir(TMP_DIR, parents=True, exist_ok=True)

# report worksheet column name
column_name = ['合約編號', '公司名稱', '合約類型', '取號狀態', '合約對象', '簽約日期', '合約年限', '合約起日', '合約訖日',
                       '建立人', '建立日期', '連結']

# email list for total report
mngr_mail_list = [
    '藍啓儒<chijulan@inaenergy.com.tw>', '李穎涵<febeli@inaenergy.com.tw>', '徐郁鈞<edwardhsu@inaenergy.com.tw>'
]

# email sender
EMAIL_SEND = '寶晶能源合約取號系統<service@inaenergy.com.tw>'


#IP and PORT
URL_ROOT = 'http://' + HOST_IP + ':' + HOST_PORT + '/'

# add excel last column url
def  add_xls_col_url(sheet: Worksheet) -> None:
    last_column = sheet.max_column
    last_row = sheet.max_row
    last_column_letter = get_column_letter(last_column)
    for n in range(1, last_row):
        cell = sheet[f'{last_column_letter}{n+1}']
        cell.hyperlink = cell.value


# User monthly check
def contract_monthly_report():
    # email to user
    users = User.objects.filter(is_active=True)
    # go through every user
    for user in users:
        print("user_email: {}".format(user.email))

        # concat 'created_by__last_name', 'created_by__first_name' to one field
        fullname = Concat('created_by__last_name', 'created_by__first_name')
        url = Concat(Value(URL_ROOT+'docnum/contract/'), 'id', Value('/'), output_field=CharField())

        # query contract data by user
        archived = Contract.history.filter(created_by=user, status__name="已歸檔", is_valid=True,
                                           update_time__month=today.month).annotate(fullname=fullname, url=url).values_list(
            'sn', 'comp__fullname','category__name', 'status__name', 'counterparty', 'sign_date', 'length', 'start_date',
            'end_date', 'fullname', 'add_time__date', 'url')

        non_archives = user.contracts.annotate(fullname=fullname, url=url).values_list(
            'sn', 'comp__fullname','category__name', 'status__name', 'counterparty', 'sign_date', 'length', 'start_date',
            'end_date', 'fullname', 'add_time__date', 'url')
        unchecks = non_archives.filter(is_valid=True, status__name="已取號")
        unachives = non_archives.filter(is_valid=True, status__name="已確認")

        # create workbook obj.
        wb = Workbook()
        ws_uncheck = wb.active
        ws_uncheck.title = "未確認"
        ws_uncheck.append(column_name)

        # write contract data into worksheet
        for row in unchecks:
            ws_uncheck.append(row)
        ws_unarchive = wb.create_sheet('未歸檔')
        ws_unarchive.append(column_name)
        for row in unachives:
            ws_unarchive.append(row)
        ws_archive = wb.create_sheet('本月已歸檔')
        ws_archive.append(column_name)
        for row in archived:
            ws_archive.append(row)

        # Due to last column is url, add hyperlink by function
        for worksheet in wb:
            # if no contact data, no need to add hyperlink.
            if worksheet.max_row != 1:
                add_xls_col_url(worksheet)

        mon_rp_file = Path(TMP_DIR, '{}{}_寶晶合約取號{}年{}月報告_{}.xlsx'.format(
            user.last_name, user.first_name, today.year, today.month, timezone.localtime(timezone.now()).strftime('%Y%m%d%H%M%S%f')))

        wb.save(mon_rp_file)

        report_summary = {
            'username': user,
            'month': today.month,
            'archived': archived.count(),
            'unarchived': unachives.count(),
            'uncheck': unchecks.count(),
        }
        if user.email != "":
            email_template = render_to_string(
                'email/monthly_report.html',
                report_summary
            )
            email = EmailMessage(
                subject='寶晶能源合約取號管理系統{}{} {}年{}月總結報告'.format(user.last_name, user.first_name, today.year, today.month),
                body=email_template,
                from_email=EMAIL_SEND,
                to=[user.email],
            )
            email.attach_file(mon_rp_file)
            email.fail_silently = False
            email_status = email.send()
            print("寄信成功") if email_status else print("寄信失敗")

        # delete data
        mon_rp_file.unlink()


def mngr_monthly_report():

    fullname = Concat('created_by__last_name', 'created_by__first_name')
    url = Concat(Value(URL_ROOT + 'docnum/contract/'), 'id', Value('/'), output_field=CharField())

    archived = Contract.history.filter(status__name="已歸檔", is_valid=True,
                                       update_time__month=today.month).annotate(fullname=fullname, url=url).values_list(
        'sn', 'comp__fullname', 'category__name', 'status__name', 'counterparty', 'sign_date', 'length', 'start_date',
        'end_date', 'fullname', 'add_time__date', 'url')

    non_archives = Contract.objects.annotate(fullname=fullname, url=url).values_list(
        'sn', 'comp__fullname', 'category__name', 'status__name', 'counterparty', 'sign_date', 'length', 'start_date',
        'end_date', 'fullname', 'add_time__date', 'url')
    unchecks = non_archives.filter(is_valid=True, status__name="已取號")
    unachives = non_archives.filter(is_valid=True, status__name="已確認")


    wb = Workbook()
    wb.active.title = '未確認'
    wb.create_sheet('未歸檔')
    wb.create_sheet('本月已歸檔')

    all_contracts = {
        '未確認': unchecks,
        '未歸檔': unachives,
        '本月已歸檔': archived,
    }

    for ws in wb:
        ws.append(column_name)
        for contra_row in all_contracts[ws.title]:
            ws.append(contra_row)

    mon_rp_file = Path(TMP_DIR, '寶晶合約取號{}年{}月總報告_{}.xlsx'.format(
        today.year, today.month, timezone.localtime(timezone.now()).strftime('%Y%m%d%H%M%S%f')))

    wb.save(mon_rp_file)

    report_summary = {
        'today': today,
        'archived': archived.count(),
        'unarchived': unachives.count(),
        'uncheck': unchecks.count(),
    }
    email_template = render_to_string(
        'email/mngr_monthly_report.html',
        report_summary
    )
    email = EmailMessage(
        subject='寶晶能源合約取號管理系統{}年{}月總結報告'.format(today.year, today.month),
        body=email_template,
        from_email=EMAIL_SEND,
        to=mngr_mail_list,
    )
    email.attach_file(mon_rp_file)
    email.fail_silently = False
    email_status = email.send()
    print("寄信成功") if email_status else print("寄信失敗")
    mon_rp_file.unlink()


def user_expired_contra_check():
    users = User.objects.filter(is_active=True)
    fullname = Concat('created_by__last_name', 'created_by__first_name')
    url = Concat(Value(URL_ROOT + 'docnum/contract/'), 'id', Value('/'), output_field=CharField())
    for user in users:
        contracts = user.docnum_contract_author.filter(is_valid=True).exclude(expiration='已終止').annotate(fullname=fullname, url=url).values_list(
            'sn', 'comp__fullname', 'category__name', 'status__name', 'counterparty', 'sign_date', 'length', 'start_date',
            'end_date', 'expiration','fullname', 'add_time__date', 'url')
        expireds = contracts.filter(end_date__lt=today)
        year_expired = contracts.filter(end_date=today.replace(year=today.year+1))
        three_mon_expired = contracts.filter(end_date=today + datetime.timedelta(days=90))
        in_mon_expired = contracts.filter(end_date__range=(today, today + datetime.timedelta(days=30)))
        expired_column_name = ['合約編號', '公司名稱', '合約類型', '取號狀態', '合約對象', '簽約日期', '合約年限', '合約起日',
                       '合約訖日','合約到期狀態', '建立人', '建立日期', '連結']
        contract_dict = {
            '已過期': expireds,
            '一個月內': in_mon_expired,
            '三個月': three_mon_expired,
            '一年': year_expired,
        }

        wb = Workbook()
        wb.active.title = '已過期'
        wb.create_sheet('一個月內')
        wb.create_sheet('三個月')
        wb.create_sheet('一年')

        for period in contract_dict:
            print(user.username, period, contract_dict[period].count())
            print(contract_dict[period])

            if contract_dict[period].count() != 0:
                wb[period].append(expired_column_name)
                for row in contract_dict[period]:
                    wb[period].append(row)
                add_xls_col_url(wb[period])
            else:
                del wb[period]
        # only email when they have expired/almost expired contacts
        if len(wb.sheetnames) != 0:
            # only email who has email
            if user.email != "":
                file = Path(TMP_DIR, '{}{}_寶晶合約到期通知.xlsx'.format(user.last_name, user.first_name))
                wb.save(file)

                summary = {
                    'expired': contract_dict['已過期'].count(),
                    'in_month': contract_dict['一個月內'].count(),
                    'three_month': contract_dict['三個月'].count(),
                    'year': contract_dict['一年'].count(),
                }

                email_template = render_to_string('email/daily_expired.html', summary)

                email = EmailMessage(
                    subject='您有合約即將到期_寶晶能源合約取號系統',
                    body=email_template,
                    from_email=EMAIL_SEND,
                    to=[user.email],
                )
                email.content_subtype = 'html'
                email.attach_file(file)
                email.fail_silently = False
                email_status = email.send()
                print("寄信成功") if email_status else print("寄信失敗")
                file.unlink()
        # delete wb object
        del wb